"""Tests for analysis export helpers."""

from __future__ import annotations

from pathlib import Path

import matplotlib
import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook
from PIL import Image

matplotlib.use("Agg")

from napari_swc_viewer.analysis.clustering import (
    ClusterRegionSelection,
    ClusterResult,
    ClusterRunMetadata,
)
from napari_swc_viewer.analysis.export import (
    export_cluster_workbook,
    export_distance_workbook,
    export_extended_parquet,
    build_clustermap_figure,
    read_extended_parquet_analysis_metadata,
    rgba_to_hex,
    save_dendrogram_figure,
)


def _write_source_parquet(path: Path) -> None:
    """Write a small neuron parquet fixture with one unclustered neuron."""
    pd.DataFrame(
        {
            "file_id": ["n1", "n1", "n2", "n2", "n3", "n3"],
            "node_id": [1, 2, 1, 2, 1, 2],
            "type": [1, 3, 1, 3, 1, 3],
            "x": [0.0, 1.0, 25.0, 26.0, 50.0, 51.0],
            "y": [0.0, 1.0, 25.0, 26.0, 50.0, 51.0],
            "z": [0.0, 1.0, 25.0, 26.0, 50.0, 51.0],
            "radius": [5.0, 1.0, 5.0, 1.0, 5.0, 1.0],
            "parent_id": [-1, 1, -1, 1, -1, 1],
            "region_id": [68, 68, 500, 500, 500, 500],
            "region_name": ["FRP1", "FRP1", "CP", "CP", "CP", "CP"],
            "region_acronym": ["FRP1", "FRP1", "CP", "CP", "CP", "CP"],
            "subject": ["s1", "s1", "s2", "s2", "s3", "s3"],
            "neuron_id": ["neuron1", "neuron1", "neuron2", "neuron2", "neuron3", "neuron3"],
        }
    ).to_parquet(path, index=False)


def _make_cluster_result(source_path: Path) -> tuple[ClusterResult, dict[str, list[float]]]:
    """Return a cluster result with export metadata."""
    result = ClusterResult(
        correlation_matrix=np.array([[1.0, 0.75], [0.75, 1.0]], dtype=np.float32),
        distance_matrix=np.array([[0.0, 0.25], [0.25, 0.0]], dtype=np.float32),
        linkage_matrix=np.array([[0.0, 1.0, 0.25, 2.0]], dtype=np.float64),
        neuron_ids=["n1", "n2"],
        reorder_indices=np.array([1, 0], dtype=np.intp),
        labels=np.array([1, 2], dtype=np.int32),
    )
    region_selection = ClusterRegionSelection(
        selected_region_ids=[184, 500],
        selected_region_acronyms=["FRP", "CP"],
        represented_region_ids=[68, 500],
        represented_region_acronyms=["FRP1", "CP"],
    )
    result.metadata = ClusterRunMetadata.from_region_selection(
        region_selection=region_selection,
        analysis_method="voxel_correlation",
        clustering_algorithm="hierarchical",
        distance_metric="one_minus_pearson_r",
        clustering_linkage="average",
        dendrogram_linkage="average",
        dilation_fraction=0.2,
        requested_cluster_count=2,
        actual_cluster_count=2,
        dbscan_eps=None,
        dbscan_min_samples=None,
        atlas_name="fake_atlas",
        atlas_resolution_um=(25.0, 25.0, 25.0),
        source_parquet_path=str(source_path),
        dendrogram_leaf_order=[1, 0],
    )
    cluster_color_map = {
        "n1": [0.12, 0.47, 0.71, 1.0],
        "n2": [0.84, 0.15, 0.16, 1.0],
    }
    return result, cluster_color_map


def _fill_rgb(cell) -> str | None:
    """Return the cell fill color as an ARGB string when present."""
    fill = cell.fill
    if fill is None:
        return None
    return getattr(fill.fgColor, "rgb", None)


def test_export_cluster_workbook_writes_colored_assignments_and_metadata(tmp_path: Path) -> None:
    source_path = tmp_path / "neurons.parquet"
    _write_source_parquet(source_path)
    result, cluster_color_map = _make_cluster_result(source_path)
    output_path = tmp_path / "clusters.xlsx"

    export_cluster_workbook(
        output_path,
        result,
        cluster_color_map,
        figure_title="Cluster Figure",
        x_label="Neuron Index",
        y_label="Neuron Index",
    )

    workbook = load_workbook(output_path)
    assert workbook.sheetnames == ["Clusters", "Metadata"]

    cluster_sheet = workbook["Clusters"]
    metadata_sheet = workbook["Metadata"]

    assert cluster_sheet["B2"].value == "n2"
    assert cluster_sheet["E2"].value == 2
    assert cluster_sheet["F2"].value == rgba_to_hex(cluster_color_map["n2"])
    assert _fill_rgb(cluster_sheet["E2"]).endswith(rgba_to_hex(cluster_color_map["n2"]).lstrip("#"))

    metadata = {
        row[0].value: row[1].value
        for row in metadata_sheet.iter_rows(min_row=2, values_only=False)
        if row[0].value
    }
    assert metadata["analysis_method"] == "voxel_correlation"
    assert metadata["figure_title"] == "Cluster Figure"
    assert metadata["source_parquet_path"] == str(source_path)


def test_export_distance_workbook_writes_dendrogram_order_and_metadata(tmp_path: Path) -> None:
    source_path = tmp_path / "neurons.parquet"
    _write_source_parquet(source_path)
    result, cluster_color_map = _make_cluster_result(source_path)
    output_path = tmp_path / "distances.xlsx"

    export_distance_workbook(
        output_path,
        result,
        cluster_color_map,
        figure_title="Distance Figure",
        x_label="Neuron A",
        y_label="Neuron B",
    )

    workbook = load_workbook(output_path)
    assert workbook.sheetnames == ["Distances", "Metadata"]

    distance_sheet = workbook["Distances"]
    metadata_sheet = workbook["Metadata"]

    assert distance_sheet["A1"].value == "Neuron ID"
    assert distance_sheet["B1"].value == "Cluster"
    assert distance_sheet["C1"].value == "n2"
    assert distance_sheet["D1"].value == "n1"
    assert distance_sheet["A2"].value == "n2"
    assert distance_sheet["B2"].value == 2
    assert float(distance_sheet["C2"].value) == 0.0
    assert float(distance_sheet["D2"].value) == 0.25
    assert _fill_rgb(distance_sheet["C1"]).endswith(rgba_to_hex(cluster_color_map["n2"]).lstrip("#"))

    metadata = {
        row[0].value: row[1].value
        for row in metadata_sheet.iter_rows(min_row=2, values_only=False)
        if row[0].value
    }
    assert metadata["figure_title"] == "Distance Figure"
    assert metadata["figure_x_label"] == "Neuron A"
    assert metadata["figure_y_label"] == "Neuron B"


def test_export_extended_parquet_preserves_rows_and_round_trips_metadata(tmp_path: Path) -> None:
    source_path = tmp_path / "neurons.parquet"
    _write_source_parquet(source_path)
    result, _cluster_color_map = _make_cluster_result(source_path)
    output_path = tmp_path / "extended.parquet"

    export_extended_parquet(output_path, result)

    loaded = pd.read_parquet(output_path).sort_values(["file_id", "node_id"]).reset_index(drop=True)
    assert len(loaded) == 6
    assert loaded.loc[loaded["file_id"] == "n1", "cluster_assignment"].dropna().unique().tolist() == [1]
    assert loaded.loc[loaded["file_id"] == "n2", "cluster_assignment"].dropna().unique().tolist() == [2]
    assert loaded.loc[loaded["file_id"] == "n3", "cluster_assignment"].isna().all()

    payload = read_extended_parquet_analysis_metadata(output_path)
    assert payload["run_metadata"]["analysis_method"] == "voxel_correlation"
    assert payload["neuron_ids_in_dendrogram_order"] == ["n2", "n1"]
    assert payload["cluster_labels_in_dendrogram_order"] == [2, 1]
    np.testing.assert_allclose(payload["distance_matrix"], result.distance_matrix)
    np.testing.assert_allclose(payload["linkage_matrix"], result.linkage_matrix)


def test_build_clustermap_figure_applies_title_and_axis_labels(tmp_path: Path) -> None:
    source_path = tmp_path / "neurons.parquet"
    _write_source_parquet(source_path)
    result, cluster_color_map = _make_cluster_result(source_path)

    figure = build_clustermap_figure(
        result,
        cluster_color_map,
        title="Publication Figure",
        x_label="X Label",
        y_label="Y Label",
    )
    try:
        assert figure._suptitle.get_text() == "Publication Figure"
        assert any(axis.get_xlabel() == "X Label" for axis in figure.axes)
        assert any(axis.get_ylabel() == "Y Label" for axis in figure.axes)
    finally:
        import matplotlib.pyplot as plt

        plt.close(figure)


def test_build_clustermap_figure_includes_distance_colorbar(tmp_path: Path) -> None:
    source_path = tmp_path / "neurons.parquet"
    _write_source_parquet(source_path)
    result, cluster_color_map = _make_cluster_result(source_path)

    figure = build_clustermap_figure(
        result,
        cluster_color_map,
    )
    try:
        assert any(
            axis.get_ylabel() == "Distance (1 - Pearson r)"
            for axis in figure.axes
        )
    finally:
        import matplotlib.pyplot as plt

        plt.close(figure)


def test_build_clustermap_figure_downsamples_large_heatmap(tmp_path: Path) -> None:
    from scipy.cluster.hierarchy import linkage

    source_path = tmp_path / "neurons.parquet"
    _write_source_parquet(source_path)
    n = 10
    result = ClusterResult(
        correlation_matrix=np.eye(n, dtype=np.float32),
        distance_matrix=np.arange(n * n, dtype=np.float32).reshape(n, n),
        linkage_matrix=linkage(np.arange(n, dtype=np.float64).reshape(-1, 1), method="single"),
        neuron_ids=[f"n{i}" for i in range(n)],
        reorder_indices=np.arange(n - 1, -1, -1, dtype=np.intp),
        labels=np.arange(1, n + 1, dtype=np.int32),
    )
    cluster_color_map = {
        neuron_id: [0.1, 0.2, 0.3, 1.0]
        for neuron_id in result.neuron_ids
    }

    figure = build_clustermap_figure(
        result,
        cluster_color_map,
        max_render_size=4,
    )
    try:
        heatmap_images = [
            axis.images[0]
            for axis in figure.axes
            if getattr(axis, "images", None)
            and axis.images
            and getattr(axis.images[0].get_array(), "ndim", 0) == 2
        ]
        assert len(heatmap_images) == 1
        assert tuple(heatmap_images[0].get_array().shape) == (4, 4)
    finally:
        import matplotlib.pyplot as plt

        plt.close(figure)


@pytest.mark.parametrize("dpi", [150, 300, 600])
def test_save_dendrogram_figure_writes_requested_dpi(tmp_path: Path, dpi: int) -> None:
    source_path = tmp_path / "neurons.parquet"
    _write_source_parquet(source_path)
    result, cluster_color_map = _make_cluster_result(source_path)
    output_path = tmp_path / f"dendrogram_{dpi}.png"

    save_dendrogram_figure(
        output_path,
        result,
        cluster_color_map,
        title="DPI Test",
        x_label="X",
        y_label="Y",
        dpi=dpi,
    )

    with Image.open(output_path) as image:
        stored_dpi = image.info.get("dpi")

    assert stored_dpi is not None
    assert round(stored_dpi[0]) == dpi
    assert round(stored_dpi[1]) == dpi
